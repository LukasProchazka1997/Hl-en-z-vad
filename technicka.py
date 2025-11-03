import csv
import tkinter as tk
from tkinter import messagebox, Toplevel, StringVar, ttk
from openpyxl import Workbook, load_workbook
import os
from datetime import datetime
import smtplib
from email.mime.text import MIMEText

# --------- KONFIGURACE -------------
CSV_FILE = "technicka.csv"
JMENA_FILE = "jmena.csv"
XLSX_FILE = "vystupts.xlsx"

# Email nastavení
SMTP_SERVER = "smtp.gmail.com"
SMTP_PORT = 587
EMAIL_USER = "vyjezdhhmh@gmail.com"
EMAIL_PASS = "lzjh kbtc xdkk jaiy"
EMAIL_TO = "lupc@post.cz"
# -----------------------------------

def nacti_csv(soubor):
    with open(soubor, newline='', encoding="utf-8") as f:
        reader = csv.reader(f)
        return [row[0] for row in reader if row]

def uloz_do_xlsx(radek, odpoved):
    cas = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    if os.path.exists(XLSX_FILE):
        wb = load_workbook(XLSX_FILE)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.append(["Původní text", "Odpověď", "Čas"])
    ws.append([radek, odpoved, cas])
    wb.save(XLSX_FILE)
    return cas

def odesli_email(radek, odpoved, cas):
    predmet = f"Hlášení k: {radek}"
    text = f"Původní text: {radek}\nOdpověď: {odpoved}\nČas: {cas}"
    msg = MIMEText(text, "plain", "utf-8")
    msg["Subject"] = predmet
    msg["From"] = EMAIL_USER
    msg["To"] = EMAIL_TO

    with smtplib.SMTP(SMTP_SERVER, SMTP_PORT) as server:
        server.starttls()
        server.login(EMAIL_USER, EMAIL_PASS)
        server.send_message(msg)

def nacti_poslednich_20():
    if not os.path.exists(XLSX_FILE):
        return []
    wb = load_workbook(XLSX_FILE)
    ws = wb.active
    zaznamy = []

    # zjistíme, ve kterém sloupci je "Řešení"
    header = [ws.cell(row=1, column=i).value for i in range(1, ws.max_column+1)]
    try:
        reseni_col = header.index("Řešení") + 1
    except ValueError:
        reseni_col = None

    for row in ws.iter_rows(min_row=2):
        radek, odpoved, cas = row[0].value, row[1].value, row[2].value
        if radek and odpoved and cas:
            text = f"[{cas}] {radek} → {odpoved}"
            if reseni_col:
                reseni = row[reseni_col-1].value
                if reseni:
                    text += f" | Řešení: {reseni}"
            zaznamy.append(text)
    return list(reversed(zaznamy[-20:]))

def vyber_jmeno_a_popis(parent):
    """Otevře okno s výběrem jména a možností napsat popis"""
    jmena = nacti_csv(JMENA_FILE)
    if not jmena:
        messagebox.showerror("Chyba", "Soubor jmena.csv je prázdný nebo neexistuje.", parent=parent)
        return None, None

    top = Toplevel(parent)
    top.title("Vyber jméno a napiš popis")

    tk.Label(top, text="Vyber jméno:").pack(pady=(5,0))
    vybrane = StringVar(value=jmena[0])
    combo = ttk.Combobox(top, textvariable=vybrane, values=jmena, state="readonly", width=30)
    combo.pack(pady=5)

    tk.Label(top, text="Popis / poznámka:").pack(pady=(5,0))
    popis_text = tk.Text(top, width=40, height=5)
    popis_text.pack(pady=5)

    odpoved = {}

    def potvrdit():
        jmeno = vybrane.get().strip()
        popis = popis_text.get("1.0", tk.END).strip()
        if not jmeno:
            messagebox.showerror("Chyba", "Je potřeba vybrat jméno.", parent=top)
            return
        if not popis:
            messagebox.showerror("Chyba", "Je potřeba napsat popis.", parent=top)
            return
        odpoved["jmeno"] = jmeno
        odpoved["popis"] = popis
        top.destroy()

    tk.Button(top, text="OK", command=potvrdit).pack(pady=10)

    top.transient(parent)
    top.grab_set()
    parent.wait_window(top)

    return odpoved.get("jmeno"), odpoved.get("popis")

# -------------------------------
# Hlavní aplikace
# -------------------------------
class Aplikace:
    def __init__(self, parent):
        """Parent je tk.Frame nebo tk.Tk"""
        self.parent = parent

        # tlačítka z CSV
        self.btn_frame = tk.Frame(parent)
        self.btn_frame.pack(pady=5)

        radky = nacti_csv(CSV_FILE)

        max_radky = 10
        max_sloupce = 6

        for idx, radek in enumerate(radky):
            r = idx // max_sloupce  # řádek
            c = idx % max_sloupce   # sloupec

            if r >= max_radky:
                break  # přesáhli jsme max počet řádků

            btn = tk.Button(self.btn_frame, text=radek, width=20, command=lambda r=radek: self.klik(r))
            btn.grid(row=r, column=c, padx=5, pady=5, sticky="nsew")

        # tlačítko pro refresh
        refresh_btn = tk.Button(parent, text="Refresh historie", width=20, command=self.refresh)
        refresh_btn.pack(pady=(5, 10))

        # rámeček pro historii
        tk.Label(parent, text="Poslední hlášení:").pack()
        # používáme Text widget pro zalamování textu
        self.history = tk.Text(parent, width=2000, height=1000, wrap="word")
        self.history.pack(pady=5)
        self.history.config(state="disabled")  # zabrání editaci uživatelem
        self.history.bind("<Double-Button-1>", self.klik_hist)

        # načtení historie při startu
        self.zpravy = nacti_poslednich_20()
        self.obnov_hist()

    def klik(self, radek):
        jmeno, popis = vyber_jmeno_a_popis(self.parent)
        if jmeno and popis:
            odpoved = f"{jmeno} → {popis}"
            try:
                cas = uloz_do_xlsx(radek, odpoved)
                odesli_email(radek, odpoved, cas)
                messagebox.showinfo("Hotovo", "Hlášení bylo uloženo a odesláno.", parent=self.parent)
                self.refresh()
            except Exception as e:
                messagebox.showerror("Chyba", f"Nastala chyba:\n{e}", parent=self.parent)

    def refresh(self):
        self.zpravy = nacti_poslednich_20()
        self.obnov_hist()

    def obnov_hist(self):
        self.history.config(state="normal")
        self.history.delete("1.0", tk.END)
        for z in self.zpravy:
            self.history.insert(tk.END, z + "\n\n")  # dvojitý enter pro oddělení záznamů
        self.history.config(state="disabled")

    # -------------------------------
    # Nová funkcionalita – zadání řešení
    # -------------------------------
    def klik_hist(self, event):
        """Po dvojkliku na záznam z historie se nabídne možnost zadat řešení po zadání hesla."""
        index = self.history.index(f"@{event.x},{event.y}")
        line_start = f"{index.split('.')[0]}.0"
        line_end = f"{int(index.split('.')[0])+1}.0"
        zaznam_text = self.history.get(line_start, line_end).strip()

        if not zaznam_text:
            return

        def overeni_hesla():
            if heslo_var.get() == "1234":
                top.destroy()
                zadat_reseni()
            else:
                messagebox.showerror("Chyba", "Špatné heslo.", parent=top)

        def zadat_reseni():
            top2 = Toplevel(self.parent)
            top2.title("Zadat řešení")

            tk.Label(top2, text="Řešení:").pack(pady=(5,0))
            reseni_text = tk.Text(top2, width=40, height=5)
            reseni_text.pack(pady=5)

            def ulozit_reseni():
                reseni = reseni_text.get("1.0", tk.END).strip()
                if not reseni:
                    messagebox.showerror("Chyba", "Je potřeba napsat řešení.", parent=top2)
                    return
                # uložíme řešení do XLSX
                self.uloz_reseni_do_xlsx(zaznam_text, reseni)
                messagebox.showinfo("Hotovo", "Řešení bylo uloženo.", parent=top2)
                top2.destroy()
                self.refresh()

            tk.Button(top2, text="Uložit", command=ulozit_reseni).pack(pady=10)
            top2.transient(self.parent)
            top2.grab_set()
            self.parent.wait_window(top2)

        top = Toplevel(self.parent)
        top.title("Ověření hesla")

        tk.Label(top, text="Zadejte heslo:").pack(pady=(5,0))
        heslo_var = StringVar()
        tk.Entry(top, textvariable=heslo_var, show="*").pack(pady=5)
        tk.Button(top, text="Potvrdit", command=overeni_hesla).pack(pady=10)

        top.transient(self.parent)
        top.grab_set()
        self.parent.wait_window(top)

    def uloz_reseni_do_xlsx(self, zaznam_text, reseni):
        """Uloží řešení do sloupce 'Řešení' odpovídající původnímu textu a odpovědi"""
        if not os.path.exists(XLSX_FILE):
            messagebox.showerror("Chyba", "Soubor XLSX neexistuje.", parent=self.parent)
            return
        wb = load_workbook(XLSX_FILE)
        ws = wb.active

        # zjistíme, ve kterém sloupci je "Řešení"
        header = [ws.cell(row=1, column=i).value for i in range(1, ws.max_column+1)]
        try:
            reseni_col = header.index("Řešení") + 1
        except ValueError:
            reseni_col = ws.max_column + 1
            ws.cell(row=1, column=reseni_col, value="Řešení")

        for row in ws.iter_rows(min_row=2):
            radek_val = row[0].value
            odpoved_val = row[1].value
            cas_val = row[2].value
            kombinace = f"[{cas_val}] {radek_val} → {odpoved_val}"
            if kombinace == zaznam_text.split(" | Řešení")[0]:  # ignorujeme existující text řešení
                ws.cell(row=row[0].row, column=reseni_col, value=reseni)
                wb.save(XLSX_FILE)
                break

# --- Spuštění aplikace ---
if __name__ == "__main__":
    root = tk.Tk()
    root.title("Hlášení")
    app = Aplikace(root)
    root.mainloop()

