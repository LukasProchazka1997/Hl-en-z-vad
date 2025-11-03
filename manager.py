# hlášení_all.py — kompletní Streamlit verze
import streamlit as st
import pandas as pd
import os
from datetime import datetime
import smtplib
from email.mime.text import MIMEText
import csv

# --------- KONFIGURACE ---------
SERVICES = {
    "Spojová služba": {"csv": "spojova.csv", "xlsx": "vystupss.xlsx"},
    "Technická služba": {"csv": "technicka.csv", "xlsx": "vystupts.xlsx"},
    "Strojní služba": {"csv": "strojni.csv", "xlsx": "vystupsts.xlsx"},
}
JMENA_FILE = "jmena.csv"

SMTP_SERVER = "smtp.gmail.com"
SMTP_PORT = 587
EMAIL_USER = "vyjezdhhmh@gmail.com"
EMAIL_PASS = "lzjh kbtc xdkk jaiy"
EMAIL_TO = "lupc@post.cz"

PASSWORD_FOR_RESENI = "1234"
# -----------------------------------

# --------- FUNKCE ---------
def nacti_csv(soubor):
    if not os.path.exists(soubor):
        return []
    with open(soubor, newline='', encoding="utf-8") as f:
        reader = csv.reader(f)
        return [row[0] for row in reader if row]

def uloz_do_xlsx(xlsx_file, radek, odpoved):
    cas = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    from openpyxl import Workbook, load_workbook
    if os.path.exists(xlsx_file):
        wb = load_workbook(xlsx_file)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.append(["Původní text", "Odpověď", "Čas"])
    ws.append([radek, odpoved, cas])
    wb.save(xlsx_file)
    return cas

def odesli_email(radek, odpoved, cas):
    predmet = f"Hlášení k: {radek}"
    text = f"Původní text: {radek}\nOdpověď: {odpoved}\nČas: {cas}"
    msg = MIMEText(text, "plain", "utf-8")
    msg["Subject"] = predmet
    msg["From"] = EMAIL_USER
    msg["To"] = EMAIL_TO
    try:
        with smtplib.SMTP(SMTP_SERVER, SMTP_PORT) as server:
            server.starttls()
            server.login(EMAIL_USER, EMAIL_PASS)
            server.send_message(msg)
    except Exception as e:
        st.warning(f"Chyba při odesílání e-mailu: {e}")

def nacti_poslednich_20(xlsx_file):
    if not os.path.exists(xlsx_file):
        return []
    from openpyxl import load_workbook
    wb = load_workbook(xlsx_file)
    ws = wb.active
    zaznamy = []
    header = [ws.cell(row=1, column=i).value for i in range(1, ws.max_column+1)]
    try:
        reseni_col = header.index("Řešení") + 1
    except ValueError:
        reseni_col = None
    for row in ws.iter_rows(min_row=2):
        radek, odpoved, cas = row[0].value, row[1].value, row[2].value
        if radek and odpoved and cas:
            text = f"[{cas}] {radek} → {odpoved}"
            if reseni_col:
                reseni = row[reseni_col-1].value
                if reseni:
                    text += f" | Řešení: {reseni}"
            zaznamy.append(text)
    return list(reversed(zaznamy[-20:]))

def uloz_reseni_do_xlsx(xlsx_file, zaznam_text, reseni):
    if not os.path.exists(xlsx_file):
        st.warning("Soubor XLSX neexistuje.")
        return
    from openpyxl import load_workbook
    wb = load_workbook(xlsx_file)
    ws = wb.active
    header = [ws.cell(row=1, column=i).value for i in range(1, ws.max_column+1)]
    try:
        reseni_col = header.index("Řešení") + 1
    except ValueError:
        reseni_col = ws.max_column + 1
        ws.cell(row=1, column=reseni_col, value="Řešení")
    for row in ws.iter_rows(min_row=2):
        radek_val = row[0].value
        odpoved_val = row[1].value
        cas_val = row[2].value
        kombinace = f"[{cas_val}] {radek_val} → {odpoved_val}"
        if kombinace == zaznam_text.split(" | Řešení")[0]:
            ws.cell(row=row[0].row, column=reseni_col, value=reseni)
            wb.save(xlsx_file)
            break

# --------- STREAMLIT ---------
st.set_page_config(page_title="Hlášení služeb", layout="wide")
st.title("Hlášení služeb")

tabs = st.tabs(list(SERVICES.keys()) + ["Manager"])

# --- Manager Tab ---
with tabs[-1]:
    st.header("CSV Manager")
    csv_file = st.file_uploader("Načíst CSV (pro Manager)", type="csv", key="manager_upload")
    if csv_file:
        df = pd.read_csv(csv_file, header=None)
        st.session_state['manager_data'] = df[0].tolist()
    if 'manager_data' not in st.session_state:
        st.session_state['manager_data'] = []

    manager_data = st.session_state['manager_data']

    edited_data = st.text_area("Uprav položky (každá položka na nový řádek):", "\n".join(manager_data), height=300)
    if st.button("Uložit CSV"):
        new_data = [line.strip() for line in edited_data.split("\n") if line.strip()]
        save_path = st.text_input("Uložit jako:", "manager.csv")
        if save_path:
            with open(save_path, "w", newline='', encoding="utf-8") as f:
                writer = csv.writer(f)
                for item in new_data:
                    writer.writerow([item])
            st.success(f"Soubor uložen: {save_path}")

# --- Service Tabs ---
for idx, (service_name, cfg) in enumerate(SERVICES.items()):
    with tabs[idx]:
        st.header(service_name)
        radky = nacti_csv(cfg['csv'])
        if not radky:
            st.warning("CSV soubor je prázdný nebo neexistuje.")
            continue
        for radek in radky:
            st.write(f"**{radek}**")
            jmena = nacti_csv(JMENA_FILE)
            if jmena:
                jmeno = st.selectbox("Vyber jméno:", jmena, key=f"{service_name}_{radek}")
            else:
                jmeno = ""
            popis = st.text_area("Popis / poznámka:", key=f"popis_{service_name}_{radek}")
            if st.button("Potvrdit hlášení", key=f"submit_{service_name}_{radek}"):
                odpoved = f"{jmeno} → {popis}"
                cas = uloz_do_xlsx(cfg['xlsx'], radek, odpoved)
                odesli_email(radek, odpoved, cas)
                st.success("Hlášení uloženo a odesláno!")

        st.subheader("Poslední hlášení")
        zaznamy = nacti_poslednich_20(cfg['xlsx'])
        for z in zaznamy:
            st.write(z)
            reseni_input = st.text_input(f"Řešení pro {z}:", key=f"reseni_{service_name}_{z}")
            heslo = st.text_input(f"Heslo pro {z}:", type="password", key=f"heslo_{service_name}_{z}")
            if st.button(f"Uložit řešení pro {z}", key=f"save_reseni_{service_name}_{z}"):
                if heslo == PASSWORD_FOR_RESENI:
                    uloz_reseni_do_xlsx(cfg['xlsx'], z, reseni_input)
                    st.success("Řešení uloženo.")
                else:
                    st.error("Špatné heslo!")
