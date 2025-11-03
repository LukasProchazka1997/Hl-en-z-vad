import streamlit as st
import pandas as pd
from openpyxl import Workbook, load_workbook
import os
from datetime import datetime
import smtplib
from email.mime.text import MIMEText
import csv

# --------- KONFIGURACE -------------
CSV_FILE = "strojni.csv"
JMENA_FILE = "jmena.csv"
XLSX_FILE = "vystupsts.xlsx"

SMTP_SERVER = "smtp.gmail.com"
SMTP_PORT = 587
EMAIL_USER = "vyjezdhhmh@gmail.com"
EMAIL_PASS = "lzjh kbtc xdkk jaiy"
EMAIL_TO = "lupc@post.cz"
# -----------------------------------

# --------- FUNKCE ---------
def nacti_csv(soubor):
    if not os.path.exists(soubor):
        return []
    with open(soubor, newline='', encoding="utf-8") as f:
        reader = csv.reader(f)
        return [row[0] for row in reader if row]

def uloz_do_xlsx(radek, odpoved):
    cas = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    if os.path.exists(XLSX_FILE):
        wb = load_workbook(XLSX_FILE)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.append(["Původní text", "Odpověď", "Čas"])
    ws.append([radek, odpoved, cas])
    wb.save(XLSX_FILE)
    return cas

def odesli_email(radek, odpoved, cas):
    predmet = f"Hlášení k: {radek}"
    text = f"Původní text: {radek}\nOdpověď: {odpoved}\nČas: {cas}"
    msg = MIMEText(text, "plain", "utf-8")
    msg["Subject"] = predmet
    msg["From"] = EMAIL_USER
    msg["To"] = EMAIL_TO
    try:
        with smtplib.SMTP(SMTP_SERVER, SMTP_PORT) as server:
            server.starttls()
            server.login(EMAIL_USER, EMAIL_PASS)
            server.send_message(msg)
    except Exception as e:
        st.warning(f"Chyba při odesílání e-mailu: {e}")

def nacti_poslednich_20():
    if not os.path.exists(XLSX_FILE):
        return []
    wb = load_workbook(XLSX_FILE)
    ws = wb.active
    zaznamy = []
    header = [ws.cell(row=1, column=i).value for i in range(1, ws.max_column+1)]
    try:
        reseni_col = header.index("Řešení") + 1
    except ValueError:
        reseni_col = None
    for row in ws.iter_rows(min_row=2):
        radek, odpoved, cas = row[0].value, row[1].value, row[2].value
        if radek and odpoved and cas:
            text = f"[{cas}] {radek} → {odpoved}"
            if reseni_col:
                reseni = row[reseni_col-1].value
                if reseni:
                    text += f" | Řešení: {reseni}"
            zaznamy.append(text)
    return list(reversed(zaznamy[-20:]))

def uloz_reseni_do_xlsx(zaznam_text, reseni):
    if not os.path.exists(XLSX_FILE):
        st.warning("Soubor XLSX neexistuje.")
        return
    wb = load_workbook(XLSX_FILE)
    ws = wb.active
    header = [ws.cell(row=1, column=i).value for i in range(1, ws.max_column+1)]
    try:
        reseni_col = header.index("Řešení") + 1
    except ValueError:
        reseni_col = ws.max_column + 1
        ws.cell(row=1, column=reseni_col, value="Řešení")

    for row in ws.iter_rows(min_row=2):
        radek_val = row[0].value
        odpoved_val = row[1].value
        cas_val = row[2].value
        kombinace = f"[{cas_val}] {radek_val} → {odpoved_val}"
        if kombinace == zaznam_text.split(" | Řešení")[0]:
            ws.cell(row=row[0].row, column=reseni_col, value=reseni)
            wb.save(XLSX_FILE)
            break

# --------- STREAMLIT APLIKACE ---------
st.set_page_config(page_title="Hlášení strojní služby", layout="wide")
st.title("Hlášení strojní služby")

tabs = st.tabs(["Strojní služba"])

with tabs[0]:
    st.subheader("Strojní služba")
    radky = nacti_csv(CSV_FILE)
    col_count = 6

    for idx, radek in enumerate(radky):
        r = idx // col_count
        c = idx % col_count
        btn_key = f"strojni_{idx}"
        if st.button(radek, key=btn_key):
            jmena = nacti_csv(JMENA_FILE)
            jmeno = st.selectbox("Vyber jméno:", jmena, key=f"jmeno_{btn_key}")
            popis = st.text_area("Popis / poznámka:", key=f"popis_{btn_key}")
            if st.button("Potvrdit", key=f"potvrdit_{btn_key}"):
                odpoved = f"{jmeno} → {popis}"
                cas = uloz_do_xlsx(radek, odpoved)
                odesli_email(radek, odpoved, cas)
                st.success("Hlášení uloženo a odesláno!")

    st.markdown("**Poslední hlášení:**")
    zaznamy = nacti_poslednich_20()
    for z in zaznamy:
        # Zadávání řešení po heslu
        if st.button(f"Zadat řešení: {z}"):
            heslo = st.text_input("Zadejte heslo:", type="password")
            if heslo == "1234":
                reseni = st.text_area("Řešení:", key=f"reseni_{z}")
                if st.button("Uložit řešení", key=f"ulozit_{z}"):
                    uloz_reseni_do_xlsx(z, reseni)
                    st.success("Řešení bylo uloženo.")
            elif heslo:
                st.error("Špatné heslo.")
